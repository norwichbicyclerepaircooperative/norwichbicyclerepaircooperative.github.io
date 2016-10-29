"""
Collate the xlsx files downloaded by the dl_invoices.py script
"""

import os
from openpyxl import Workbook, load_workbook

wb_out = Workbook()
sheet_out = wb_out.active
sheet_out.append(("invoice_no", "invoice_date", "product_code", "product_desc", "qty", "your_price", "total_price"))

for invoice_no in os.listdir("./invoices/"):
    assert invoice_no.isdigit()
    filename = os.path.join("invoices", invoice_no, "Invoice.xlsx")
    assert os.path.isfile(filename)
    book = load_workbook(filename)
    sheet = book.active
    invoice_no =sheet['K3'].value
    invoice_date = sheet['J5'].value
    # item code; item price, number purchased & total cost
    assert sheet.cell(column=1, row=18).value == "Product Code"
    row = 19
    while True:
        product_code = sheet.cell(column=1, row=row).value
        if not product_code:
            break
        product_desc = sheet.cell(column=7, row=row).value
        qty = sheet.cell(column=19, row=row).value
        your_price = sheet.cell(column=25, row=row).value
        total_price = sheet.cell(column=26, row=row).value
        print(invoice_no, invoice_date, product_code, product_desc, qty, your_price, total_price)
        assert round(qty * your_price, 2) == total_price
        row += 1
        sheet_out.append((invoice_no, invoice_date, product_code, product_desc, qty, your_price, total_price))

wb_out.save('combined_invoices.xlsx')
