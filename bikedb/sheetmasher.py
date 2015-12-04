#!/usr/bin/python3
"""
A thing to combine the co-op's three input spreadsheets into a unified output
spreadsheet for analysis etc.
"""

from openpyxl import Workbook, load_workbook
import os

working_dir = os.path.dirname(os.path.abspath(__file__))


class NoSuchNumberException(Exception):
    pass


class BikeSheet(object):
    """Helper class for easy access to spreadsheet cells by col name or
    by bike number"""

    # Disgusting hack because the header row isn't always the first row
    header_rows = {
        'Bicycles in.xlsx': 2,
        'Bicycles completed.xlsx': 1,
        'Bicycle sold.xlsx': 1,
    }

    def __init__(self, filename):
        self.header_row = self.header_rows[filename]
        filename = os.path.join(working_dir, filename)
        self.book = load_workbook(filename)
        self.sheet = self.book.active  # book.get_sheet_by_name("Sheet1")
        self.get_cols()

    def get_cols(self):
        """ Make a dict of column names and positions"""
        self.cols = {}
        for col in range(1, 100):
            v = self.sheet.cell(row=self.header_row, column=col)
            if v.value is None:
                continue
            self.cols[v.value] = col
            col += 1

    def get_val_by_row(self, col_name, row):
        return self.sheet.cell(row=row, column=self.cols[col_name]).value

    def get_val_by_number(self, number, col_name):
        for row in range(2, 99999):
            v = self.get_val_by_row("Number", row)
            if v is None:
                raise NoSuchNumberException
            if v == number:
                return self.get_val_by_row(col_name, row)

    def get_vals_by_row(self, row):
        vals = {}
        for col_name, col_no in self.cols.items():
            vals[col_name] = self.sheet.cell(row=row, column=col_no).value
        return vals

    def get_vals_by_number(self, number):
        for row in range(2, 99999):
            v = self.get_val_by_row("Number", row)
            if v is None:
                return dict([(f, None) for f in self.cols.keys()])
            if v == number:
                vals = {}
                for col_name, col_no in self.cols.items():
                    vals[col_name] = self.sheet.cell(row=row, column=col_no).value
                return vals


def process():
    bikes_in = BikeSheet("Bicycles in.xlsx")
    bikes_completed = BikeSheet("Bicycles completed.xlsx")
    bikes_sold = BikeSheet("Bicycle sold.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(("Number", "Code", "Make", "Model", "Frame", "Type", "Colour",
               "Frame number", "Date acquired", "Destination",
               "Serial No. or Invoice No.", "Purchase price", "Date finished",
               "Mechanic", "Date sold", "Receipt no.", "Sale price",))
    for bi_row in range(3, 99999):
        bifields = bikes_in.get_vals_by_row(bi_row)
        number = bifields['Number']
        if number is None:
            # Assume we're done if we find a row where the Number field is blank
            print("Done reading, saving spreadsheet...")
            break
        bcfields = bikes_completed.get_vals_by_number(number)
        bsfields = bikes_sold.get_vals_by_number(number)

        out_row = (
            number,
            bifields['Code'],
            bifields['Make'],
            bifields['Model'],
            bcfields['Frame'],
            bcfields['Type'],
            bifields['Colour'],
            bcfields['Frame_number'],
            bifields['Date_acquired'],
            bifields['Destination'],
            bifields['Serial No. or Invoice No.'],
            bifields['Price_paid'],
            bcfields['Date_finished'],
            bcfields['Mechanic'],
            bsfields['Date_sold'],
            bsfields['Receipt-no.'],
            bsfields['Price_sold'],
        )
        #print(out_row)
        ws.append(out_row)
    output_file = os.path.join(working_dir, "out.xlsx")
    wb.save(output_file)


if __name__ == "__main__":
    process()
